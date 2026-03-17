import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.config import OUTPUT_DIR

log = logging.getLogger(__name__)

_PURPLE      = "4B2E83"
_LIGHT_LILAC = "EDE7F6"
_WHITE       = "FFFFFF"

_HEADERS = ["Dia",  "Data",  "Hora início", "Hora fim", "Local", "Descrição", "Preço"]
_WIDTHS  = [15,      13,      13,              13,          35,      45,       15      ]

_DAY_LABEL = {
    "sexta":   "Sexta-feira",
    "sábado":  "Sábado",
    "domingo": "Domingo",
}


def _end_time(start) -> str:
    """Return end time = start + 4h, wrapping past midnight."""
    if not start or start == "—":
        return "—"
    try:
        h, m = map(int, start.split(":"))
        end_h = (h + 4) % 24
        return f"{end_h:02d}:{m:02d}"
    except Exception:
        return "—"


def load_events_from_excel(path: Path) -> list[dict]:
    """
    Read events back from an existing agenda Excel file.
    Returns a list of dicts with keys: day_of_week, date, time, time_end, location, description, price.
    """
    _DAY_LABEL_REVERSE = {v: k for k, v in _DAY_LABEL.items()}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = (row + (None,) * 7)[:7]
        day_pt, date_val, time_start, time_end, location, description, price = cells
        if not date_val or not location:  # skip empty / timestamp rows
            continue
        events.append({
            "day_of_week": _DAY_LABEL_REVERSE.get(str(day_pt or ""), str(day_pt or "")),
            "date":        str(date_val),
            "time":        None if time_start in (None, "—", "") else str(time_start),
            "time_end":    None if time_end   in (None, "—", "") else str(time_end),
            "location":    str(location),
            "description": str(description or ""),
            "price":       str(price or ""),
        })
    return events


def get_excel_path_for_today() -> Path:
    """Return the output path for today's Excel without creating it."""
    return OUTPUT_DIR / f"forro_agenda_{date.today().isoformat()}.xlsx"


def export_to_excel(events: list[dict]) -> Path:
    """
    Write events to output/forro_agenda_YYYY-MM-DD.xlsx and return the path.
    Includes an extraction timestamp row at the bottom.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dest = get_excel_path_for_today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agenda Forró"

    # ── Header row ────────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color=_WHITE, size=11)
    hdr_fill  = PatternFill("solid", fgColor=_PURPLE)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, label in enumerate(_HEADERS, start=1):
        cell            = ws.cell(row=1, column=col, value=label)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.alignment  = hdr_align

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=_LIGHT_LILAC)

    for i, ev in enumerate(events):
        row  = i + 2
        fill = alt_fill if i % 2 == 0 else None

        day_raw   = (ev.get("day_of_week") or "").lower()
        start     = ev.get("time") or None
        row_data  = [
            _DAY_LABEL.get(day_raw, ev.get("day_of_week", "")),
            ev.get("date", ""),
            start or "—",
            ev.get("time_end") or _end_time(start),
            ev.get("location", ""),
            ev.get("description", ""),
            ev.get("price", ""),
        ]

        for col, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
            if fill:
                cell.fill = fill

    # ── Extraction timestamp row ──────────────────────────────────────────────
    ts_row  = len(events) + 3
    ts_cell = ws.cell(row=ts_row, column=1,
                      value=f"Extraído em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ts_cell.font      = Font(italic=True, color="888888", size=9)
    ts_cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=ts_row, start_column=1, end_row=ts_row, end_column=6)

    # ── Column widths & freeze ────────────────────────────────────────────────
    for col, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"

    wb.save(dest)
    log.info(f"Excel exportado: {dest.name}")
    return dest
